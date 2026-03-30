"""Excel builder — Generate a 2-sheet XLSX from transaction data.

Sheet 1 "Transactions": One row per transaction with ~35 AI-extracted fields.
Sheet 2 "Autres articles": Non-transaction articles (title, summary, link).

Returns base64-encoded XLSX for transmission via JSON (Power Automate saves to SharePoint).
"""

import base64
import io
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Newmark navy for headers
NAVY_FILL = PatternFill(start_color="002855", end_color="002855", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

# Column mapping: (header_label, json_key, width)
# json_key=None means auto-generated (reference, year, etc.)
TRANSACTION_COLUMNS = [
    ("N° Réf", None, 8),
    ("Année", None, 8),
    ("Semestre", None, 8),
    ("Trimestre", None, 10),
    ("Nom immeuble", "nom_immeuble", 25),
    ("Adresse", "adresse", 30),
    ("Département / CP", "departement_cp", 14),
    ("Commune", "commune", 18),
    ("Sous-secteur géo", "sous_secteur_geo", 18),
    ("Région", "region", 16),
    ("% Bureaux", "bureaux_pct", 10),
    ("% Commerces", "commerces_pct", 10),
    ("% Industriel", "industriel_pct", 10),
    ("% Résidentiel", "residentiel_pct", 10),
    ("% Hôtel", "hotel_pct", 10),
    ("% Santé", "sante_pct", 10),
    ("Vendeur(s)", "vendeurs", 25),
    ("Profil vendeur", "profil_vendeur", 18),
    ("Nationalité vendeur", "nationalite_vendeur", 16),
    ("Acquéreur(s)", "acquereurs", 25),
    ("Profil acquéreur", "profil_acquereur", 18),
    ("Nationalité acquéreur", "nationalite_acquereur", 16),
    ("Nb acquéreurs", "nb_acquereurs", 12),
    ("Prix HD (M€)", "prix_hd", 12),
    ("Prix AEM (M€)", "prix_aem", 12),
    ("Prix retenu (M€)", "prix_retenu", 14),
    ("Estimé", "estime", 8),
    ("Surface (m²)", "surface_m2", 12),
    ("Prix / m²", "prix_m2", 10),
    ("Taux rendement (%)", "taux_rendement", 14),
    ("État", "etat", 12),
    ("Type transaction", "type_transaction", 18),
    ("VEFA", "vefa", 8),
    ("Locataire(s)", "locataires", 25),
    ("Loyer (M€)", "loyer", 10),
    ("WALB (ans)", "walb", 10),
    ("Broker(s)", "brokers", 20),
    ("Commentaires", "commentaires", 35),
    ("Contenu article", "contenu_article", 60),
    ("Source", "source", 18),
    ("URL", "url", 40),
]

OTHER_COLUMNS = [
    ("Titre", "titre", 40),
    ("Résumé", "resume", 60),
    ("Contenu article", "contenu_article", 60),
    ("Source", "source", 18),
    ("URL", "url", 40),
]


def _apply_header_style(ws, columns):
    """Write styled header row and set column widths."""
    for col_idx, (label, _key, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = NAVY_FILL
        cell.font = WHITE_FONT
        cell.alignment = WRAP_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def _get_quarter_info(dt: datetime) -> tuple[int, int, int]:
    """Return (year, semester, trimester) for a date."""
    year = dt.year
    month = dt.month
    semester = 1 if month <= 6 else 2
    trimester = (month - 1) // 3 + 1
    return year, semester, trimester


def build_transaction_excel(
    transactions: list[dict],
    other_articles: list[dict],
) -> str:
    """Build XLSX workbook and return as base64 string.

    Args:
        transactions: List of dicts from Claude (est_transaction=true).
        other_articles: List of dicts from Claude (est_transaction=false).

    Returns:
        Base64-encoded XLSX string.
    """
    wb = Workbook()
    now = datetime.now()
    year, semester, trimester = _get_quarter_info(now)

    # ── Sheet 1: Transactions ──
    ws_tx = wb.active
    ws_tx.title = "Transactions"
    _apply_header_style(ws_tx, TRANSACTION_COLUMNS)

    for row_idx, tx in enumerate(transactions, start=2):
        for col_idx, (_label, key, _width) in enumerate(TRANSACTION_COLUMNS, start=1):
            if key is None:
                # Auto-generated columns
                if _label == "N° Réf":
                    value = row_idx - 1
                elif _label == "Année":
                    value = year
                elif _label == "Semestre":
                    value = f"S{semester}"
                elif _label == "Trimestre":
                    value = f"T{trimester}"
                else:
                    value = None
            else:
                value = tx.get(key)

            cell = ws_tx.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP_ALIGN

    # ── Sheet 2: Autres articles ──
    ws_other = wb.create_sheet("Autres articles")
    _apply_header_style(ws_other, OTHER_COLUMNS)

    for row_idx, art in enumerate(other_articles, start=2):
        for col_idx, (_label, key, _width) in enumerate(OTHER_COLUMNS, start=1):
            value = art.get(key)
            cell = ws_other.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP_ALIGN

    # ── Encode to base64 ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    b64 = base64.b64encode(buffer.read()).decode("utf-8")

    logging.info(
        "Excel generated: %d transactions, %d other articles",
        len(transactions),
        len(other_articles),
    )
    return b64
