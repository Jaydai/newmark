"""Excel report generator for property owner lookup results.

Produces a 2-sheet XLSX workbook:
  Sheet 1 — "Propriétaires"  : one row per address with owner details
  Sheet 2 — "Dirigeants"     : one row per dirigeant from La Place company pages

Returns the workbook as a base64-encoded string for JSON transport.
"""

import base64
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Newmark branding
NAVY_FILL = PatternFill(start_color="002855", end_color="002855", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")

# Sheet 1 columns: (header, width)
OWNER_COLUMNS = [
    ("Adresse", 40),
    ("Propriétaire", 30),
    ("Type", 18),
    ("SIREN", 15),
    ("Forme juridique", 22),
    ("Maison mère", 25),
    ("Nbre actifs", 12),
    ("Surface (m²)", 14),
    ("Erreurs", 40),
]

# Sheet 2 columns
DIRIGEANT_COLUMNS = [
    ("Adresse", 40),
    ("Propriétaire", 30),
    ("Fonction", 18),
    ("Nom", 25),
    ("Société du contact", 30),
]


def build_report(results: list[dict], dirigeants: dict[str, list[dict]]) -> str:
    """Build the XLSX report and return as base64.

    Args:
        results: List of per-address result dicts from the pipeline.
        dirigeants: Dict mapping address → list of dirigeant dicts from La Place.
    """
    wb = Workbook()

    # --- Sheet 1: Propriétaires ---
    ws1 = wb.active
    ws1.title = "Propriétaires"
    _write_header(ws1, OWNER_COLUMNS)

    for row_idx, r in enumerate(results, start=2):
        values = [
            r.get("address", ""),
            r.get("owner_name", ""),
            r.get("owner_type", ""),
            r.get("siren", ""),
            r.get("forme_juridique", ""),
            r.get("maison_mere", ""),
            r.get("nbre_actifs", ""),
            r.get("surface", ""),
            r.get("error", "") or "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP_ALIGN

    # --- Sheet 2: Dirigeants ---
    ws2 = wb.create_sheet("Dirigeants")
    _write_header(ws2, DIRIGEANT_COLUMNS)

    row_idx = 2
    for r in results:
        address = r.get("address", "")
        owner_name = r.get("owner_name", "")
        for d in dirigeants.get(address, []):
            values = [
                address,
                owner_name,
                d.get("fonction", ""),
                d.get("nom", ""),
                d.get("societe", ""),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = WRAP_ALIGN
            row_idx += 1

    # Encode to base64
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")


def _write_header(ws, columns: list[tuple]):
    """Write styled header row."""
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = NAVY_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[cell.column_letter].width = width
